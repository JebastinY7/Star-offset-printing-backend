import json
import random
import time
from decimal import Decimal
from django.core.mail import send_mail
from django.contrib.auth.models import User
from .models import PasswordResetOTP
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from .models import Customer, Bill, BillItem, OffersHistory, Setting, MembershipTransaction, LoginAttempt, PointTransaction, Order
from django.utils import timezone
from .utils import send_whatsapp_template
from datetime import timedelta, datetime, date
from openpyxl.styles import Font, Alignment
from django.http import JsonResponse
from django.http import HttpResponse
from openpyxl import Workbook
from django.utils.timezone import localtime
from django.db.models import Q, Sum
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth, TruncYear
from openpyxl.utils import get_column_letter
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from pricing.models import Category, Size, PriceRule, CategoryDiscount, Variant

# Create your views here.
# Admin Login

MAX_ATTEMPTS = 5
BLOCK_MINUTES = 10

def admin_login(request):
    if request.user.is_authenticated:
        return redirect('/dashboard/')

    if request.method == "POST":
        email = request.POST.get("email").lower().strip()
        password = request.POST.get("password")

        if not email or not password:
            return render(request, 'login.html', {
                'error': 'Email and password required'
            })

        obj, _ = LoginAttempt.objects.get_or_create(email=email)

        block_time = timezone.now() - timedelta(minutes=BLOCK_MINUTES)

        if obj.attempts >= MAX_ATTEMPTS and obj.last_attempt > block_time:
            return render(request, 'login.html', {
                'error': 'Too many attempts. Try again later'
            })

        user = authenticate(request, username=email, password=password)

        if user:
            obj.attempts = 0
            obj.save()

            login(request, user)

            if not request.POST.get('remember'):
                request.session.set_expiry(0)

            return redirect('/dashboard/')
        else:
             obj.attempts += 1
             obj.save()

             return render(request, 'login.html', {'error': 'Invalid email or password'})
    
    return render(request, 'login.html')

# Forgot Password
def forgot_password(request):
    if request.method == "POST":
        email = request.POST.get("email").lower().strip()

        if not email:
            return render(request, 'forgot.html', {
                'error': 'Please enter your email'
            })

        try:
            user = User.objects.get(username=email)
        except User.DoesNotExist:
            return render(request, 'forgot.html', {
                'success': 'Enter a registered email'
            })
        
        last_otp = PasswordResetOTP.objects.filter(user=user).last()
        if last_otp and timezone.now() - last_otp.created_at < timedelta(seconds=30):
            return render(request, 'forgot.html', {
                'error': 'Please wait before requesting OTP again'
            })
        
        PasswordResetOTP.objects.filter(user=user).delete()
        
        otp = str(random.randint(100000, 999999))
        PasswordResetOTP.objects.create(user=user, otp=otp)

        send_mail (
            'Your OTP Code',
            f'Your OTP is {otp}',
            'staroffsets@gmail.com',
            [email],
            fail_silently=False
        )

        request.session['reset_email'] = email

        return redirect('verify_otp')
    
    return render(request, 'forgot.html')

# OTP Verification

MAX_OTP_ATTEMPTS = 5

def verify_otp(request):
    if request.method == "POST":
        otp = request.POST.get("otp")
        email = request.session.get('reset_email')

        if not email:
            return redirect('/')
        
        try:
            user = User.objects.get(username=email)
        except User.DoesNotExist:
            return redirect('/')
        
        valid_otp = PasswordResetOTP.objects.filter(user=user).last()

        if not valid_otp:
            return render(request, 'verify_otp.html', {
                'error': 'OTP expired. Request new one'
            })
        
        if valid_otp.otp != otp:
            valid_otp.attempts += 1
            valid_otp.save()

            if valid_otp.attempts >= MAX_OTP_ATTEMPTS:
                valid_otp.delete()
                return render(request, 'verify_otp.html', {
                    'error': 'Too many wrong attempts. Request new OTP'
                })
            
            return render(request, 'verify_otp.html', {
                'error': f'Invalid OTP. {MAX_OTP_ATTEMPTS - valid_otp.attempts} attempts left'
            })
        
        valid_otp.delete()
        return redirect('reset_password')
    
    return render(request, 'verify_otp.html')


# Reset Password
def reset_password(request):
    email = request.session.get('reset_email')

    if not email:
        return redirect('/')
    
    try:
        user = User.objects.get(username=email)
    except User.DoesNotExist:
        return redirect('/')
    
    if request.method == "POST":
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        if not password or not confirm_password:
            return render(request, 'reset_password.html', {
                'error': 'All fields are required'
            })
        
        if password != confirm_password:
            return render(request, 'reset_password.html', {
                'error': 'Passwords do not match'
            })
        
        if len(password) < 6:
            return render(request, 'reset_password.html', {
                'error': 'Password must be at least 6 characters'
            })
        
        user.set_password(password)
        user.save()

        if 'reset_email' in request.session:
            del request.session['reset_email']
        
        messages.success(request, "Password updated successfully. Please login")

        return redirect('/')
    
    return render(request, 'reset_password.html')

# Re-send OTP
def resend_otp(request):
    email = request.session.get('reset_email')

    if not email:
        return redirect('forgot_password')
    
    try:
        user = User.objects.get(username=email)
    except User.DoesNotExist:
        return redirect('forgot_password')
    
    last_otp = PasswordResetOTP.objects.filter(user=user).last()
    if last_otp and timezone.now() - last_otp.created_at < timedelta(seconds=30):
        return render(request, 'verify_otp.html', {
            'errorr': 'Please wait 30 seconds before requesting a new OTP'
        })
    
    PasswordResetOTP.objects.filter(user=user).delete()

    otp = str(random.randint(100000, 999999))
    PasswordResetOTP.objects.create(user=user, otp=otp)

    send_mail(
        'Your OTP Code',
        f'Your new OTP is {otp}',
        'shijujs1212@gmail.com',
        [email],
        fail_silently=False
    )

    return redirect('verify_otp')

# Logout
def admin_logout(request):
    logout(request)
    request.session.flush()
    return redirect('login')

# Dashboard
@never_cache
@login_required(login_url='/')
def dashboard(request):
    today = timezone.now().date()

    total_customers = Customer.objects.count()
    students = Customer.objects.filter(category="Student").count()
    customers = Customer.objects.filter(category="Customer").count()
    shops = Customer.objects.filter(category="Shop").count()

    active_cards = Customer.objects.filter(
        is_member = True,
        expiry_date__gte=today
    ).count()

    expired_cards = Customer.objects.filter(
        is_member = True,
        expiry_date__lt=today
    ).count()

    no_card = Customer.objects.filter(is_member=False).count()

    today = localtime().date()

    today_bills = Bill.objects.filter(bill_date=today)
    today_sales = today_bills.aggregate(Sum('final_amount'))['final_amount__sum'] or 0
    today_bill_count = today_bills.count()
    today_discount = today_bills.aggregate(Sum('discount'))['discount__sum'] or 0


    pending_count = Order.objects.filter(status="pending").count()
    progress_count = Order.objects.filter(
        status__in=["progress", "completed"]
    ).exclude(status="delivered").count()

    completed_count = Order.objects.filter(status="delivered").count()

    overdue_count = Order.objects.filter(
        delivery_date__lt=today
    ).exclude(status="delivered").count()

    # today_points = PointTransaction.objects.filter(
    #     type='redeem',
    #     date__date=today
    # ).aggregate(Sum('points_used'))['points_used__sum'] or 0

    today_due = Bill.objects.filter(
        bill_date=today,
        due_amount__gt = 0
    ).aggregate(total=Sum("due_amount"))["total"] or 0
    
    recent_bills = Bill.objects.select_related('customer').order_by('-id')[:5]

    recent_customers = Customer.objects.order_by('-id')[:5]

    upcoming = today + timedelta(days = 5)

    expiring_customers = Customer.objects.filter(
        is_member = True,
        expiry_date__range=[today, today+timedelta(days=5)]
    )

    recent_due_customers = Customer.objects.filter(
        due_amount__gt=0
    ).order_by('-id')[:5]

    for c in expiring_customers:
        c.days_left = (c.expiry_date - today).days

    return render(request, 'dashboard.html', {
        'total_customers': total_customers,
        'students': students,
        'customers': customers,
        'shops': shops,

        'active_cards': active_cards,
        'expired_cards': expired_cards,
        'no_card': no_card,

        'today_sales': today_sales,
        'today_bill_count': today_bill_count,
        'today_discount': today_discount,
        'today_due': today_due,

        'pending_count': pending_count,
        'progress_count': progress_count,
        'completed_count': completed_count,
        'overdue_count': overdue_count,

        'recent_bills': recent_bills,
        'recent_customers': recent_customers,
        'expiring_customers': expiring_customers,
        'recent_due_customers': recent_due_customers,
    })

def customers_list(request):
    customers = Customer.objects.all().order_by('-id')
    settings = Setting.objects.first()

    setting = Setting.objects.first()

    search = request.GET.get("search") or ""
    category = request.GET.get("category") or ""
    user_type = request.GET.get("type") or ""

    if search:
        customers = customers.filter(
           Q(name__icontains=search) |
           Q(phone__icontains=search)
        )

    if category and category != "All":
        customers = customers.filter(category=category)
    
    if user_type == "member":
        customers = customers.filter(is_member=True)
    elif user_type == "non":
        customers = customers.filter(is_member=False)
    
    query_params = request.GET.copy()

    if 'page' in query_params:
        query_params.pop('page')

    query_string = query_params.urlencode()

    paginator = Paginator(customers, 8)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    

    return render(request, 'customers.html', {
        'page_obj': page_obj,
        'today': timezone.now().date(),
        'default_fee': settings.renewal_fee if settings else 0,
        'search': search,
        'category': category,
        'user_type': user_type,
        'query_string': query_string,
        "setting": setting,
    })

def add_customer(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        phone = request.POST.get("phone", "").strip()
        category = request.POST.get("category")
        is_member = request.POST.get("is_member") == "on"
        member_id = request.POST.get("member_id", "").strip()

        join_date = request.POST.get("join_date")
        expiry_date = request.POST.get("expiry_date")

        if not name:
            messages.error(request, "Name is required")
            return render(request, 'add_customer.html')

        if not phone:
            messages.error(request, "Phone number is required")
            return render(request, 'add_customer.html')

        if Customer.objects.filter(phone=phone).exists():
            messages.error(request, "Phone number already exists")
            return render(request, 'add_customer.html')

        if is_member:
            if not member_id:
                messages.error(request, "Membership ID is required")
                return render(request, 'add_customer.html')
            
            if Customer.objects.filter(member_id=member_id).exists():
                messages.error(request, "Membership ID already exists")
                return render(request, 'add_customer.html')
    
            try:
                join_date = datetime.strptime(join_date, "%d-%m-%Y").date()
                expiry_date = datetime.strptime(expiry_date, "%d-%m-%Y").date()
            except:
                messages.error(request, "Invalid date format")
                return render(request, 'add_customer.html')
        else:
            member_id = None
            join_date = None
            expiry_date = None

        Customer.objects.create(
            name=name,
            phone=phone,
            category=category,
            is_member=is_member,
            member_id=member_id,
            join_date=join_date,
            expiry_date=expiry_date
        )

        messages.success(request, "Customer added successfully")
        return redirect('/customers/')
    
    return render(request, 'add_customer.html')

def edit_customer(request, id):
    customer = get_object_or_404(Customer, id=id)

    if request.method == "POST":
        name = request.POST.get("name")
        phone = request.POST.get("phone")
        category = request.POST.get("category")
        is_member = request.POST.get("is_member") == "on"
        member_id = request.POST.get("member_id")
        join_date = request.POST.get("join_date")
        expiry_date = request.POST.get("expiry_date")

        customer.name = name
        customer.phone = phone
        customer.category = category
        customer.is_member = is_member

        if Customer.objects.exclude(id=customer.id).filter(phone=phone).exists():
            messages.error(request, "Phone already exists")
            return render(request, 'edit_customer.html', {'customer': customer})

        if is_member:
            if not member_id:
                messages.error(request, "Membership ID is required")
                return render(request, 'edit_customer.html', {'customer': customer})
            
            if Customer.objects.exclude(id=customer.id).filter(member_id=member_id).exists():
                messages.error(request, "Membership ID already exists")
                return render(request, 'edit_customer.html', {'customer': customer})
            
            try:
                customer.join_date = datetime.strptime(join_date, "%d-%m-%Y").date()
                customer.expiry_date = datetime.strptime(expiry_date, "%d-%m-%Y").date()
            except:
                messages.error(request, "Invalid date format")
                return render(request, 'edit_customer.html', {'customer': customer})
            
            customer.member_id = member_id
        else:
            customer.member_id = None
            customer.join_date = None
            customer.expiry_date = None

        customer.save()

        messages.success(request, "Customer updated successfully")
        return redirect('/customers/')
    
    return render(request, 'edit_customer.html', {'customer': customer})

@require_POST
def delete_customer(request, id):
    customer = get_object_or_404(Customer, id=id)
    customer.delete()
    messages.success(request, "Customer deleted successfully")
    return redirect('/customers/')


def search_customer(request):
    phone = request.GET.get("phone")
    print("input:", phone)

    try:
        customer = Customer.objects.get(phone=phone)
        print("found:", customer.name)

        return JsonResponse  ({
            "id": customer.id,
            "name": customer.name,
            "category": customer.category,
            "is_member": customer.is_member,
            "member_id": customer.member_id,
            "due_amount": float(customer.due_amount),
            "points": customer.points
        })
        
    except Customer.DoesNotExist:
        print("not found")
        return JsonResponse({"error": "not found"})

def live_search_customers(request):
    query = request.GET.get("q", "")

    customers = Customer.objects.filter(
        Q(name__icontains=query) |
        Q(phone__startswith=query)
    )[:5]
    

    data = [
        {
            "id": c.id,
            "name": c.name,
            "phone": c.phone,
            "is_member": c.is_member,
            "member_id": c.member_id,
            "points": c.points,
            "category": c.category,
            "due_amount": float(c.due_amount),
            "expiry_date": c.expiry_date.strftime("%Y-%m-%d") if c.expiry_date else None
        }
        for c in customers
    ]

    return JsonResponse({"customers": data})

def invoice(request, bill_id):
    bill = get_object_or_404(Bill, id=bill_id)
    items = BillItem.objects.filter(bill=bill)

    return render(request, 'invoice.html', {
        'bill': bill,
        'items': items,
    })

def billing_system(request):
    bill_id = request.GET.get("bill_id")

    bill = None
    items = []

    categories = Category.objects.all()

    if bill_id:
        bill = get_object_or_404(Bill, id=bill_id)
        items = BillItem.objects.filter(bill=bill)
        

    return render(request, 'billing.html', {
        'bill': bill,
        'items': items,
        'categories': categories
    })

def bill_history(request):
    query = request.GET.get('q')

    bills = Bill.objects.select_related('customer').order_by('-bill_date') #-id

    if query:
        bills = bills.filter(
            Q(customer__name__icontains=query) |
            Q(bill_no__icontains=query)
        )
    
    query_params = request.GET.copy()

    if 'page' in query_params:
        query_params.pop('page')

    query_string = query_params.urlencode()
    
    paginator = Paginator(bills, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'bill_history.html', {
        'page_obj': page_obj,
        'query_string':query_string
    })

@require_POST
def delete_bill(request, id):
    bill = get_object_or_404(Bill, id=id)
    bill.delete()
    messages.success(request, "Bill deleted successfully")
    return redirect('/history/')


def reports(request):
    bills = Bill.objects.select_related('customer')

    
    category = request.GET.get('category')
    member_type = request.GET.get("member_type")
    date_range = request.GET.get("date_range")

    today = timezone.now().date()

    start = None
    end = None

    if date_range:
        dates = date_range.split(" to ")
    
        if len(dates) == 2:
            start = datetime.strptime(dates[0].strip(), "%Y-%m-%d").date()
            end = datetime.strptime(dates[1].strip(), "%Y-%m-%d").date()
        else:
            start = datetime.strptime(dates[0].strip(), "%Y-%m-%d").date()
            end = start

            

    else:
        start = today - timedelta(days=6)
        end = today


    bills = bills.filter(bill_date__range=[start, end])

    if category and category != "all":
        bills = bills.filter(customer__category=category)

    
    if member_type == "member":
        bills = bills.filter(customer__is_member=True)
    
    elif member_type == "non_member":
        bills = bills.filter(customer__is_member=False)

    diff = (end - start).days

    if diff <= 7:
        group = TruncDate('bill_date')
        label_format = "%d %b"

    elif diff <= 31:
        group = TruncDate('bill_date')
        label_format = "%d %b"

    elif diff <= 365:
        group = TruncMonth('bill_date')
        label_format = "%d %Y"

    else:
        group = TruncDate('bill_date')
        label_format = "%b %Y"


    total_sales = bills.aggregate(Sum('final_amount'))['final_amount__sum'] or 0
    total_bills = bills.count()
    total_customers = bills.values('customer').distinct().count()
    total_discount = bills.aggregate(Sum('discount'))['discount__sum'] or 0
    total_due = bills.aggregate(Sum('due_amount'))['due_amount__sum'] or 0


    daily_data = (
        bills
        .values('bill_date')
        .annotate(total=Sum('final_amount'))
        .order_by('bill_date')
    )

    highest_day = daily_data.order_by('-total').first()
    lowest_day = daily_data.order_by('total').first()


    # chart_data = (
    #     bills
    #     .annotate(period=group)
    #     .values('period')
    #     .annotate(total=Sum('final_amount'))
    #     .order_by('period')
    # )

    # labels = [item['period'].strftime(label_format) for item in chart_data if item['period']]
    # data = [float(item['total']) for item in chart_data]

    chart_data = (
    bills
    .annotate(period=group)
    .values('period')
    .annotate(total=Sum('final_amount'))
    .order_by('period')
)

    from collections import OrderedDict
    chart_map = {
        item["period"]: float(item["total"])
        for item in chart_data
        if item["period"]
    }

    labels = []
    data = []

    if diff <= 31:
        current = start
        while current <= end:
            labels.append(current.strftime(label_format))
            data.append(chart_map.get(current, 0))
            current += timedelta(days=1)
    
    elif diff <= 365:
        current = date(start.year, start.month, 1)
        while current <= end:
            labels.append(current.strftime("%b %Y"))
            data.append(chart_map.get(current, 0))

            if current.month == 12:
                current = date(current.year + 1, 1, 1)
            else:
                current = date(current.year, current.month + 1, 1)
    
    else:
        current = start
        while current <= end:
            labels.append(current.strftime(label_format))
            data.append(chart_map.get(current, 0))
            current += timedelta(days=30)

    top_customer = (
        bills
        .values('customer__name')
        .annotate(total=Sum('final_amount'))
        .order_by('-total')
        .first()
    )
    bills = bills.order_by('-bill_date')

    paginator = Paginator(bills, 14)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')

    query_string = query_params.urlencode() 

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'partials/report_table.html', {
            'page_obj': page_obj,
            'query_params': query_params,
    })

    return render(request, 'reports.html', {
        'page_obj': page_obj,
        'total_sales': total_sales,
        'total_bills': total_bills,
        'total_customers': total_customers,
        'total_due': total_due,
        'total_discount': total_discount,
        'chart_labels': json.dumps(labels),
        'chart_data': json.dumps(data),
        'highest_day': highest_day,
        'lowest_day': lowest_day,
        'top_customer': top_customer,
        'start_date': start,
        'end_date': end,
        'query_params': query_string,
    })

#Membership
def delete_membership(request, id):
    membership = get_object_or_404(MembershipTransaction, id=id)

    if request.method == "POST":
        membership.customer.is_member = False
        membership.customer.save()

        membership.delete()

        messages.success(request, "Membership deleted and customer membership removed successfully")
    
    return redirect('membership_history')


def offers(request):
    if request.method == "POST":
        title = request.POST.get("title")
        message_text = request.POST.get("message")
        group = request.POST.get("group", "all")
        category = request.POST.get("category", "all")

        if not title or not message_text:
            messages.error(request, "Title and Message required")
            return redirect('offers')

        customers = Customer.objects.all()

        if group == "members":
            customers = customers.filter(is_member=True)
        elif group == "non_members":
            customers = customers.filter(is_member=False)

        if category != "all":
            category_map = {
                "student": "Student",
                "customer": "Customer",
                "shop": "Shop",
            }
            if category in category_map:
                customers = customers.filter(
                    category__iexact=category_map[category]
                )

        customers = customers.exclude(phone__isnull=True).exclude(phone="")

        success = 0
        failed = 0
        
        # count = customers.count()

        clean_message = f"{title} - {message_text}".replace("\n", " ").strip()
        shop_number = request.POST.get("shop_number")

        for customer in customers:
            try:
                phone = customer.phone

                response = send_whatsapp_template(
                    phone=phone,
                    customer_name=customer.name,
                    offer_message=clean_message,
                    shop_number=shop_number
                )

                time.sleep(1)

                if response.get("messages"):
                    success += 1
                else:
                    failed += 1
                    print("Whatsapp failed", response)

            except Exception as e:
                print("Error:", e)
                failed += 1

        OffersHistory.objects.create(
            title=title,
            message=message_text,
            group=group,
            category=category,
            total_sent=success
        )

        messages.success(
            request, 
            f"Offer sent ✅ Success: {success}, Failed: {failed} "
        )

        return redirect('offers')

    return render(request, 'offers.html')



def offer_history(request):
    query = request.GET.get("q")

    offers = OffersHistory.objects.all().order_by('-id')

    if query:
        offers = offers.filter(
            Q(title__icontains=query) 
        )
    
    query_params = request.GET.copy()

    if 'page' in query_params:
        query_params.pop('page')

    query_string = query_params.urlencode()

    paginator = Paginator(offers, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'offer_history.html', { 
        'page_obj': page_obj,
        'query_string': query_string
    })

@require_POST
def delete_offer(request, id):
    offer = get_object_or_404(OffersHistory, id=id)
    offer.delete()
    messages.success(request, "Offer deleted successfully")
    return redirect('offer_history')

def settings(request):
    setting = Setting.objects.first()

    if not setting:
        setting = Setting.objects.create(
            # points_per_rupee = 1,
            # max_redeem_percent = 50,
            membership_validity_days = 365,
            renewal_fee = 0,
            shop_renewal_fee = 0
        )

    if request.method == "POST":
        form_type = request.POST.get("form_type")

        if form_type == "points":
            setting.points_per_rupee = request.POST.get("points_per_rupee")
            setting.max_redeem_percent = request.POST.get("max_redeem_percent")

        elif form_type == "membership":
            setting.membership_validity_days = request.POST.get("membership_validity_days")
            setting.renewal_fee = request.POST.get("renewal_fee")
            setting.shop_renewal_fee = request.POST.get("shop_renewal_fee")

        setting.save()
        messages.success(request, "Settings updated successfully")
        return redirect('settings')
    
    return render(request, 'settings.html', { 'setting': setting })

def custom_round_amount(amount):
    whole = int(amount)
    decimal = amount - whole

    if decimal < 0.5:
        return whole

    return whole + 1

def save_bill(request):
    if request.method == "POST":

        bill_id = request.POST.get('bill_id')
        

        customer_id = request.POST.get("customer_id")
        items_data = request.POST.get("items_data")

        if not customer_id:
            messages.error(request, "Customer not selected")
            return redirect('/billing/')
        
        if not items_data:
            messages.error(request, "No items added")
            return redirect('/billing/')

        items = json.loads(items_data)

        if len(items) == 0:
            messages.error(request, "No items added")
            return redirect('/billing/')

        customer = get_object_or_404(Customer, id=customer_id)

        today = timezone.now().date()

        customer_type = "Customer"

        if customer.category.lower() == "shop":
            customer_type = "Shop"
        
        elif (
            customer.category.lower() != "shop" and
            customer.is_member and
            customer.expiry_date and
            customer.expiry_date >= today
        ):
            customer_type = "Member"

        gross_total = sum(Decimal(str(item['price'])) * Decimal(str(item['qty'])) for item in items)

        item_discount_total = sum(Decimal(str(item['discount'])) for item in items)
        extra_discount = Decimal(request.POST.get("extraDiscount") or 0)

        if customer_type == "Shop":
            total_discount = item_discount_total + extra_discount
        
        elif customer_type == "Member":
            total_discount = item_discount_total + extra_discount
        
        else:
            total_discount = extra_discount

        # points = int(float(request.POST.get("points") or 0))if customer_type == "Member" else 0

        if customer_type == "Member" and customer.due_amount <=0:
            points = 0
        else:
            points = 0

        total_extra_charge = sum(Decimal(str(item.get("extraCharge", 0))) for item in items)

        raw_final = max(gross_total - total_discount + total_extra_charge - Decimal(points), Decimal('0'))

        final = Decimal(custom_round_amount(raw_final))

        paid_raw = request.POST.get("paid_amount", "").strip()

        if paid_raw == "":
            paid_amount =Decimal('0')
        else:
            paid_amount = Decimal(paid_raw)

        old_due_payment = Decimal(request.POST.get("old_due_payment") or 0)
        old_due_payment = min(old_due_payment, customer.due_amount)

        previous_due = customer.due_amount

        current_due = max(final - paid_amount, Decimal('0'))

        # current_bill_due = max(final - paid_amount, Decimal('0'))

        remaining_old_due = max(previous_due - old_due_payment, Decimal('0'))

        due_amount = current_due + remaining_old_due

        customer.due_amount = due_amount
        customer.save()

        if paid_amount >= final:
            payment_status = "paid"
        elif paid_amount > 0:
            payment_status = "partial"
        else:
            payment_status = "due"

        if bill_id:
            bill = get_object_or_404(Bill, id=int(bill_id))

            customer.due_amount = max(customer.due_amount - bill.due_amount + bill.old_due_paid, Decimal('0'))

            old_transactions = PointTransaction.objects.filter(bill=bill)

            for t in old_transactions:
                if t.type == 'redeem':
                    customer.points += t.points_used
                elif t.type == 'earn':
                    customer.points -= t.points_added

            old_transactions.delete()


            bill.customer = customer
            bill.gross_total = gross_total
            bill.extra_charge_total=total_extra_charge
            bill.total_amount = gross_total - total_discount + total_extra_charge
            bill.discount = total_discount
            bill.points_used = points
            bill.final_amount = final
            bill.paid_amount = paid_amount
            bill.old_due_paid = old_due_payment
            bill.due_amount = due_amount
            bill.previous_due = customer.due_amount
            bill.payment_status = payment_status
            # bill.bill_date = timezone.now()
            bill.save()

            print("OLD:", customer.due_amount)

            # delete old items
            bill.billitem_set.all().delete()
        
        else:
            last_bill = Bill.objects.order_by('-id').first()
            next_id = last_bill.id + 1 if last_bill else 1
        
            


            bill = Bill.objects.create(
                bill_no=f"B{next_id}",
                customer=customer,
                gross_total=gross_total,
                extra_charge_total=total_extra_charge,
                total_amount=gross_total - total_discount + total_extra_charge,
                discount=total_discount,
                points_used=points,
                paid_amount = paid_amount,
                due_amount = due_amount,
                previous_due=previous_due,
                current_due=current_due,
                payment_status = payment_status,
                final_amount=final,
                old_due_paid=old_due_payment,
                bill_date=timezone.now()
            )

        # due_amount = remaining_old_due + current_due

        

        # customer.due_amount = max(customer.due_amount - old_due_payment, Decimal('0'))

        for item in items:
            BillItem.objects.create(
                bill=bill,
                service_name=item['name'],
                category_id=item.get('categoryId') or None,
                size_id=item.get('sizeId') or None,
                variant_id=item.get('variantId') or None,
                qty=item['qty'],
                price=item['price'],
                discount=item['discount'],
                extra_charge=item.get("extraCharge", 0),
                extra_purpose=item.get("extraPurpose", ""),
                total=item['total'],
            )

        # customer.points = max(customer.points - points, 0)

        today = timezone.now().date()

        is_active_member = (customer_type == "Member")

        if is_active_member and points > 0:
            customer.points -= points

            PointTransaction.objects.create(
                customer=customer,
                bill=bill,
                points_used=int(points),
                type='redeem'
            )
        else:
            points = 0

        if is_active_member and due_amount <= 0 and customer.due_amount <= 0:
            earned_points = 0

            customer.points += earned_points

            if earned_points > 0:
                PointTransaction.objects.create(
                    customer=customer,
                    bill=bill,
                    points_added=earned_points,
                    type='earn'
                )
        else:
            earned_points = 0

        customer.save()

        return redirect(f'/invoice/{bill.id}/')

    return redirect('/billing/')

# Due
def clear_due(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)

    if request.method == "POST":
        amount = Decimal(request.POST.get("amount") or 0)

        customer.due_amount = max(customer.due_amount - amount, Decimal('0'))
        customer.save()

        messages.success(request, "Due updated successfully")
    
    return redirect('/customers/')

def renew_membership(request, id):
    customer = Customer.objects.get(id=id)

    if request.method == "POST":
        payment_method = request.POST.get("payment_method")

        settings = Setting.objects.first()
        days = settings.membership_validity_days

        if customer.category and customer.category.strip().lower() == "shop":
            amount = settings.shop_renewal_fee
        else:
            amount = settings.renewal_fe

            # if customer.category.strip().lower() == "shop":
            #     customer.category = "Customer"

        start = timezone.now().date()
        end = start + timedelta(days=days)

        customer.is_member = True
        customer.join_date = start
        customer.expiry_date = end
        customer.save()

        MembershipTransaction.objects.create(
            customer=customer,
            type='renewal',
            amount = amount,
            payment_method = payment_method,
            start_date = start,
            end_date = end
        )

    return redirect('membership_history')



# Download Excel
def download_report(request):
    bills = Bill.objects.select_related('customer')

    category = request.GET.get('category')
    member_type = request.GET.get('member_type')
    date_range = request.GET.get("date_range")

    today = timezone.now().date()

    start = None
    end = None

    if date_range:
        dates = date_range.split(" to ")

        if len(dates) == 2:
            start = datetime.strptime(dates[0].strip(), "%Y-%m-%d").date()
            end = datetime.strptime(dates[1].strip(), "%Y-%m-%d").date()

    if not start and not end:
        start = today - timedelta(days=6)
        end = today

    bills = bills.filter(bill_date__range=[start, end])

    if category and category != "all":
        bills = bills.filter(customer__category=category)

    if member_type == "member":
        bills = bills.filter(customer__is_member=True)
    elif member_type == "non_member":
        bills = bills.filter(customer__is_member=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    ws.merge_cells("A1:H1")
    ws["A1"] = "STAR OFFSET PRINTING"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = center

    ws.merge_cells("A3:H3")
    if start and end:
        ws["A3"] = f"From {start} To {end}"
    else:
        ws["A3"] = "All Data"

    ws["A3"].alignment = center

    ws.append([
        "Bill No", "Date", "Customer Name", "Category",
        "Total Amount", "Discount","Due", "Final Amount"
    ])

    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = bold

    total_sales = 0
    total_discount = 0
    total_due = 0


    for bill in bills:
        ws.append([
            bill.bill_no,
            bill.bill_date.strftime("%d-%m-%Y"),
            bill.customer.name,
            bill.customer.category,
            float(bill.total_amount),
            float(bill.discount),
            float(bill.due_amount),
            float(bill.final_amount)
        ])

        total_sales += float(bill.final_amount)
        total_discount += float(bill.discount)
        total_due += float(bill.due_amount)

    #Total Row
    ws.append([])
    ws.append([
        "", "", "", "TOTAL", "",
        total_discount,
        total_due,
        total_sales
    ])

    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = bold

    for col in ws.iter_cols():
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + 2

    reponse = HttpResponse(
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    reponse['Content-Disposition'] = 'attachment; filename=report.xlsx'

    wb.save(reponse)
    return reponse
    

# Membership Renewal History
def membership_history(request):
    transactions = MembershipTransaction.objects.select_related('customer').order_by('-created_at')

    start = request.GET.get('start')
    end = request.GET.get('end')

    if start and end:
        try:
            start_date = datetime.strptime(start, "%Y-%m-%d").date()
            end_date = datetime.strptime(end, "%Y-%m-%d").date()

            transactions = transactions.filter(
                created_at__date__range=(start_date, end_date)
            )
        except:
            pass
    
    total = transactions.aggregate(Sum('amount'))['amount__sum'] or 0

    paginator = Paginator(transactions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'membership_history.html', {
        'page_obj': page_obj,
        'total': total,
        'start': start,
        'end': end
    })

def update_password(request):
    if request.method == "POST":
        current = request.POST.get("current_password")
        new = request.POST.get("new_password")
        confirm = request.POST.get("confirm_password")

        user = request.user

        # 1️⃣ Check current password
        if not user.check_password(current):
            messages.error(request, "Current password is incorrect")
            return redirect('settings')

        # 2️⃣ Check new match
        if new != confirm:
            messages.error(request, "New passwords do not match")
            return redirect('settings')

        # 3️⃣ Optional: validation
        if len(new) < 6:
            messages.error(request, "Password must be at least 6 characters")
            return redirect('settings')

        # 4️⃣ Set new password
        user.set_password(new)
        user.save()

        messages.success(request, "Password updated successfully")
        return redirect('login')  # force re-login

    return redirect('settings')

# From Pricing App
# Auto size show-up

def get_sizes(request):
    category_id = request.GET.get("category_id")

    sizes = Size.objects.filter(category_id=category_id)

    data = {
        "sizes": [
            {
                "id": size.id,
                "name": size.name
            }
            for size in sizes
        ]
    }

    return JsonResponse(data)


def get_price(request):
    category_id = request.GET.get("category_id")
    size_id = request.GET.get("size_id")
    qty = request.GET.get("qty")
    variant_id = request.GET.get("variant_id")

    # Required fields
    if not (category_id and size_id and qty and variant_id):
        return JsonResponse({
            "price": 0,
            "shop_discount": 0,
            "cs_discount": 0,
            "notes": ""
        })

    try:
        qty = int(qty)
    except (ValueError, TypeError):
        return JsonResponse({
            "price": 0,
            "shop_discount": 0,
            "cs_discount": 0,
            "notes": ""
        })

    # Match exact category + size + variant + qty slab
    price_rule = PriceRule.objects.filter(
        category_id=category_id,
        size_id=size_id,
        variant_id=variant_id,
        min_qty__lte=qty
    ).filter(
        Q(max_qty__gte=qty) | Q(max_qty__isnull=True)
    ).order_by("min_qty").first()

    if not price_rule:
        return JsonResponse({
            "price": 0,
            "shop_discount": 0,
            "cs_discount": 0,
            "notes": ""
        })

    # Main price
    

    return JsonResponse({
        "price": float(price_rule.price or 0),
        "shop_discount": float(price_rule.shop_discount or 0),
        "cs_discount": float(price_rule.cs_discount or 0),
        "notes": price_rule.notes or ""
    })

def get_discount(request):
    category_id = request.GET.get("category_id")
    customer_type = request.GET.get("customer_type")

    if not (category_id and customer_type):
        return JsonResponse({"discount":0})
    
    discount = CategoryDiscount.objects.filter(
        category_id=category_id,
        member_type__name__iexact=customer_type
    ).first()

    if discount:
        return JsonResponse({
            "discount": discount.discount_percent
        })
    
    return JsonResponse({
        "discount": 0
    })

def get_variants(request):
    size_id = request.GET.get("size_id")

    variants = Variant.objects.filter(
        size_id=size_id
    ).order_by("display_order", "name")

    return JsonResponse({
        "variants": [
            {
                "id": v.id,
                "name": v.name
            }
            for v in variants
        ]
    })


# Orders
def orders_page(request):
    query = request.GET.get("q", "")
    status_filter = request.GET.get("status", "")

    orders_list = Order.objects.select_related("customer").order_by("-id")

    # SEARCH
    if query:
        orders_list = orders_list.filter(
            Q(customer__name__icontains=query) |
            Q(work_name__icontains=query)
        )

    # STATUS FILTER
    if status_filter:
        orders_list = orders_list.filter(status=status_filter)

    # PAGINATION
    paginator = Paginator(orders_list, 7)  # 10 orders per page
    page_number = request.GET.get("page")
    orders = paginator.get_page(page_number)

    today = timezone.now().date()

    # SUMMARY COUNTS
    pending_count = Order.objects.filter(status="pending").count()
    progress_count = Order.objects.filter(status="progress").count()
    completed_count = Order.objects.filter(status="completed").count()

    overdue_count = Order.objects.filter(
        delivery_date__lt=today
    ).exclude(status__in=["delivered", "cancelled"]).count()

    return render(request, "orders.html", {
        "orders": orders,
        "today": today,
        "pending_count": pending_count,
        "progress_count": progress_count,
        "completed_count": completed_count,
        "overdue_count": overdue_count,
    })


# Update Order
def update_order_status(request, id, new_status):
    order = get_object_or_404(Order, id=id)

    valid_flow = {
        "pending": "progress",
        "progress": "completed",
        "completed": "delivered",
    }

    current_status = order.status


    if current_status in valid_flow and valid_flow[current_status] == new_status:
        order.status = new_status
        order.save()

    return redirect('/orders/')

# Edit Order
def edit_order(request, id):
    order = get_object_or_404(Order, id=id)

    if request.method == "POST":
        order.work_name = request.POST.get("work_name")
        order.notes = request.POST.get("notes")

        order.order_date = request.POST.get("order_date")
        order.delivery_date = request.POST.get("delivery_date")

        order.total_amount = Decimal(request.POST.get("total_amount") or 0)
        order.advance_paid = Decimal(request.POST.get("advance_paid") or 0)

        order.status = request.POST.get("status")

        customer_id = request.POST.get("customer_id")
        order.customer = get_object_or_404(Customer, id=customer_id)

        order.save()

        messages.success(request, "Order updated successfully")
        return redirect('/orders/')

    customers = Customer.objects.all().order_by('name')

    return render(request, 'edit_order.html', {
        'order': order,
        'customers': customers
    })

def add_order(request):
    customers = Customer.objects.all().order_by("name")

    if request.method == "POST":
        
        customer_id = request.POST.get("customer")
        if not customer_id:
            messages.error(request, "Please select a customer")
            return render(request, "add_order.html", {
                "customers": customers
            })
        
        work_name = request.POST.get("work_name")
        notes = request.POST.get("notes")

        order_date = request.POST.get("order_date")
        delivery_date = request.POST.get("delivery_date")

        total_amount = Decimal(request.POST.get("total_amount") or 0)
        advance_paid = Decimal(request.POST.get("advance_paid") or 0)

        customer = get_object_or_404(Customer, id=customer_id)

        if total_amount > 0:
            due_amount = max(total_amount - advance_paid, 0)
        else:
            due_amount = Decimal(0)

        if advance_paid < 0:
            advance_paid = Decimal(0)
        
        if advance_paid > total_amount and total_amount > 0:
            advance_paid = total_amount

        Order.objects.create(
            customer=customer,
            work_name=work_name,
            notes=notes,
            order_date=order_date,
            delivery_date=delivery_date,
            total_amount=total_amount,
            advance_paid=advance_paid,
            due_amount=due_amount,
            status="pending"
        )

        messages.success(request, "Order added successfully")
        return redirect("orders_page")

    else:
        return render(request,"add_order.html",
            {"customers": customers,
             "today": date.today()}
        )

# Delete Order
@require_POST
def delete_order(request, id):
    order = get_object_or_404(Order, id=id)
    order.delete()

    messages.success(request, "Order deleted successfully")
    return redirect('/orders/')